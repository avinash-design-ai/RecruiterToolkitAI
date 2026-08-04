import os

from openpyxl import load_workbook

from config import INPUT_FOLDER
from occupations import OCCUPATION_MAP


class ExcelReader:

    def __init__(self, file_path=None):

        if file_path:

            self.input_file = file_path

        else:

            files = [

                f for f in os.listdir(INPUT_FOLDER)

                if f.endswith(".xlsx")
                and not f.startswith("~$")

            ]

            if not files:

                raise Exception(
                    "No Excel file found."
                )

            self.input_file = os.path.join(
                INPUT_FOLDER,
                files[0]
            )

        self.workbook = load_workbook(
            self.input_file
        )

        self.sheet = self.workbook.active

    # ---------------------------------------------

    def get_headers(self):

        headers = {}

        for col in range(
            3,
            self.sheet.max_column + 1
        ):

            value = self.sheet.cell(
                2,
                col
            ).value

            if value in OCCUPATION_MAP:

                headers[col] = OCCUPATION_MAP[value]

        return headers

    # ---------------------------------------------

    def get_rows(self):

        rows = []

        row = 3

        while True:

            city = self.sheet.cell(
                row,
                1
            ).value

            county = self.sheet.cell(
                row,
                2
            ).value

            if city is None:

                break

            city = str(city).strip()

            if city == "":

                break

            if county is None:

                county = ""

            rows.append({

                "row": row,

                "city": city,

                "county": str(county).strip()

            })

            row += 1

        print(f"Total Locations : {len(rows)}")

        return rows

    # ---------------------------------------------

    def write(self, row, column, value):

        self.sheet.cell(

            row=row,

            column=column

        ).value = value

    # ---------------------------------------------

    def save(self, output_path):

        self.workbook.save(output_path)

        return output_path
