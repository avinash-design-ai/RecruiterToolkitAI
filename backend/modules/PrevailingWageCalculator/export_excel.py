from openpyxl import load_workbook


class ExcelManager:

    def __init__(self, excel_path):

        self.workbook = load_workbook(excel_path)

        self.sheet = self.workbook.active

    def total_rows(self):

        return self.sheet.max_row - 2

    def get_rows(self):

        rows = []

        for row in range(3, self.sheet.max_row + 1):

            city = self.sheet[f"A{row}"].value
            county = self.sheet[f"B{row}"].value

            if city is None:
                continue

            rows.append(
                {
                    "row": row,
                    "city": str(city).strip(),
                    "county": str(county).strip()
                }
            )

        return rows

    def write(self, row, column, value):

        # Remove commas and US$ before writing
        if isinstance(value, str):
            value = value.replace("US$", "")
            value = value.replace(",", "")
            value = value.strip()

        self.sheet[f"{column}{row}"] = value

    def save(self, output_path):

        self.workbook.save(output_path)